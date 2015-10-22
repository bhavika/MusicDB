import os
from openpyxl import Workbook



music = os.chdir('./Music');
# print os.getcwd()

#create an excel workbook


wb = Workbook()
#get current active worksheet
ws = wb.active
ws['A1'] = 'Artist'
ws['B1'] = 'Album'
ws['C1'] = 'Songs'
  



def getartists(path):
    seen = set()
    artists = []
    for root, dirs, files in os.walk(path, topdown=False):
        if dirs:
            parent = root
            while parent:
                seen.add(parent)
                parent = os.path.dirname(parent)
        for d in dirs:
            d = os.path.join(root, d)
            if d not in seen:
              artists.insert(len(artists), d)
    yield artists

def writeartists():
    for artists in getartists('.'):
        start_row = 3
        for artist in artists:
            start_column = 1
            ws.cell(row=start_row, column=start_column).value = str(artist)
            start_row += 1
            


#album inside artist
def getalbums(path):
    seen = set()
    albums = []
    for root, dirs, files in os.walk(path, topdown = False):
        if dirs:
            parent = root
            while parent: 
                seen.add(parent)
                parent = os.path.dirname(parent)
        for albms in dirs:
            albms = os.path.join(root, albums)
            if albms not in seen:
                albums.insert(len(albums), albms);
    yield albms
    
def writealbums():
    for albms in getalbums('./..'):
        start_row = 3
        for a in albms:
            print "Album name: ", a
            start_column = 2
            ws.cell(row=start_row, column=start_column).value = str(a)
            start_row += 1
  

#songs inside album


def getsongs(path):
    seen = set()
    songs = []
    for root, dirs, files in os.walk(path, topdown=False):
        if dirs:
            parent = root
            while parent:
                seen.add(parent)
                parent = os.path.dirname(parent)
        for f in files:
            f = os.path.join(f)
            if f not in seen:
                 songs.insert(len(songs), f)
    yield songs        
   
  
#writing song list
def writesongs():
    for songs in getsongs('.'):
        start_row = 3
        for track in songs:
            if track.endswith(('.mp3', '.mp4', '.flac')):
                print "Track name: ", track
                start_column = 3
                ws.cell(row=start_row, column=start_column).value = str(track)
                start_row += 1
                
                
# def MakeDB(path):
#     getartists(path)
#     getalbums('.')
#     getsongs(path)
    

getalbums(os.getcwd())


# MakeDB(music)

wb.save("MusicDB.xlsx");      
