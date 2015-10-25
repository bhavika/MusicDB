__author__ = 'bhavika'

import os
from openpyxl import Workbook

music = os.chdir('/media/bhavika/Bhavika/BHAVIKA/Music');

wb = Workbook()
ws = wb.active

ws['A1'] = 'Artist'
ws['B1'] = 'Album'
ws['C1'] = 'Songs'




def getalbums(path):
    seen = set()
    albums = []
    artists = []
    songs = []
    start_artist_row = 3
    start_track_row = 3
    for root, dirs, files in os.walk(path, topdown=False):
            if dirs:
                parent = root
                while parent:
                    seen.add(parent)
                    parent = os.path.dirname(parent)
                    artists.insert(len(artists), parent)
                    ws.cell(row=start_artist_row, column=1).value = str(parent)
                    start_artist_row += 1

                    for d in dirs:
                        d = os.path.join(d)
                        if d not in seen:
                            albums.insert(len(albums), d)
    yield albums


def writealbums():
    for albums in getalbums('.'):
        start_album_row = 3
        start_track_row = 3
        # start_artist_row = 3
        # for a in artists:
        #     start_column = 1
        #     ws.cell(row=start_artist_row, column=start_column).value = str(a)
        #     start_artist_row += 1
        for al in albums:
            start_column = 2
            ws.cell(row=start_album_row, column=start_column).value = str(al)
            start_album_row += 1



if __name__ == "__main__":
    print("Main function")
    getalbums(music)
    writealbums()
    wb.save("/home/bhavika/MusicDb.xlsx")



